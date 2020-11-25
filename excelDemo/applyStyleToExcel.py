import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.utils import get_column_letter

import CellStyles


def apply_styles_to_excel(file_path):
    wb = openpyxl.load_workbook(file_path)

    CellStyles.add_styles_to_workbook(wb)

    sheet = wb.active

    for column in sheet.columns:
        print(column)
        print(get_column_letter(column[1].column))
        length = max(len(str(cell.value)) for cell in column)
        sheet.column_dimensions[get_column_letter(column[1].column)].width = float(length * 2.5)

    for row in sheet.iter_rows(min_row=2, min_col=1):
        for cell in row:
            cell.style = "cell"
            if cell.value == "Fail":
                cell.style = "fail_cell"
            if cell.value == "Pass":
                cell.style = "pass_cell"

    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=1, max_col=4):
        for cell in row:
            cell.style = "headers"

    wb.save(file_path)
